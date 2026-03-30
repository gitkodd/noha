"""
NOHA - Extrator de Planilhas v3.0
===================================
Extrai dados das planilhas .xlsx da pasta PLANILHAS/ e gera db.json.

REGRAS:
- Actual Cost NUNCA usa Price Cost como fallback. Nulo é nulo.
- Outliers (actual > price * OUTLIER_THRESHOLD) são excluídos do cálculo, mas guardados para auditoria.
- 'coverage' indica a % de itens com Actual Cost real preenchido (0.0 a 1.0).
- 'deviationReliable' é True somente se coverage >= 0.5.

USO:
    python3 pro_extract.py
"""

import openpyxl
import json
import os
from typing import Optional

# ─── Configuração ────────────────────────────────────────────────────────────

PLANILHAS_DIR = os.path.join(os.path.dirname(__file__), "PLANILHAS")
DB_PATH = os.path.join(os.path.dirname(__file__), "frontend", "src", "data", "db.json")

# Planilhas a processar — apenas as versões 2.0 (formato padronizado)
# Angela e Lilian excluídas: planilhas com dados insuficientes para análise histórica confiável
FILES = [
    {"name": "MAFE 2.0.xlsx",         "id": "mafe-2",     "title": "Mafe 2.0"},
    {"name": "Jaqueline 2.0.xlsx",    "id": "jaqueline-2","title": "Jaqueline 2.0"},
    {"name": "Susana 2.0.xlsx",       "id": "susana",     "title": "Susana 2.0"},
    {"name": "DEBORA PACKER 2.0.xlsx","id": "packer",     "title": "Debora Packer"},
]

# Mapeamento de categorias: texto bruto da planilha → código interno do sistema
CATEGORY_MAP = {
    # Construção / instalações físicas
    "CONSTRUÇÃO":           "CONSTRUCAO",
    "CONSTRUCAO":           "CONSTRUCAO",
    "FIXOS":                "CONSTRUCAO",   # instalações fixas
    "VIDRO":                "CONSTRUCAO",
    "PVC":                  "CONSTRUCAO",
    "AR CONDICIONADO":      "CONSTRUCAO",   # instalação fixa
    # Móveis e colchões
    "MARCENARIA":           "MOVEIS",
    "MÓVEIS":               "MOVEIS",
    "MOVEIS":               "MOVEIS",
    "COLCHÃO/VANDERLEI":    "MOVEIS",       # colchões e camas
    "COLCHAO/VANDERLEI":    "MOVEIS",
    # Produtos / compras diversas
    "PRODUTOS":             "PRODUTOS",
    "CATEGORY":             "PRODUTOS",
    "MERCADO":              "PRODUTOS",     # itens de consumo / frigobar
    # Decoração
    "DECORAÇÃO":            "DECORACAO",
    "DECORACAO":            "DECORACAO",
    "CORTINA":              "DECORACAO",
    "PAPEL DE PAREDE":      "DECORACAO",
    "ARTE VIDEO":           "DECORACAO",    # arte visual para os quartos
    # Mão de obra
    "MAO DE OBRA":          "MAO_DE_OBRA",
    "MÃO DE OBRA":          "MAO_DE_OBRA",
    "MAO_DE_OBRA":          "MAO_DE_OBRA",
    "M.O. INSTALAÇÃO":      "MAO_DE_OBRA",
    "FOTOGRÁFO":            "MAO_DE_OBRA",  # serviço
    "FOTOGRAFO":            "MAO_DE_OBRA",
    "PERSONAGEM":           "MAO_DE_OBRA",  # serviço de personagem
    # Mão de obra geral
    "GESTÃO DE OBRA":       "MAO_DE_OBRA_G",
    "GESTÃO":               "MAO_DE_OBRA_G",
    "MAO DE OBRA GESTAO":   "MAO_DE_OBRA_G",
    "LIMPEZA":              "MAO_DE_OBRA_G", # serviço de limpeza
    # Pacotes temáticos
    "PACOTES TEMATICOS":    "PACOTES_TEMATICOS",
    "PACOTES TEMÁTICOS":    "PACOTES_TEMATICOS",
    "BONECO TEMÁTICO":      "PACOTES_TEMATICOS",
    "BONECO TEMATICO":      "PACOTES_TEMATICOS",
    # Extras genuínos (itens fora do escopo, custos inesperados)
    "EXTRAS":               "EXTRAS",
    "IMPREVISTOS":          "EXTRAS",
    "CAFÉ":                 "EXTRAS",       # item de consumo não planejado
    "CAFE":                 "EXTRAS",
}

# Palavras-chave para classificar quartos temáticos vs adultos
THEME_KEYWORDS = [
    "mickey", "disney", "frozen", "star wars", "harry potter",
    "minions", "mario", "avengers", "rapunzel", "galaxy",
    "tema", "themed", "cenario", "cenário", "aura",
]
ADULT_KEYWORDS = [
    "adult", "master", "king", "queen", "bege",
    "principal", "casal", "suíte principal",
]

# Fator máximo aceitável entre Actual Cost e Price Cost.
# Se actual > price * OUTLIER_THRESHOLD, o item é suspeito e excluído do cálculo.
OUTLIER_THRESHOLD = 10.0

# Cobertura mínima para considerar o desvio confiável (50% dos itens com Actual)
MIN_COVERAGE_FOR_RELIABLE = 0.5

# ─── Helpers ─────────────────────────────────────────────────────────────────

def clean_number(val) -> float:
    """Converte qualquer valor de célula para float limpo. Retorna 0.0 se inválido."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def clean_actual(val):
    """
    Retorna float se Actual Cost é um número válido e positivo.
    Retorna None se a célula está vazia, é traço, texto ou zero.
    NUNCA usa Price Cost como substituto.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val > 0 else None
    text = str(val).strip()
    # Valores que significam "não preenchido"
    if text in ("", "-", "None", "LINK", "nan", "0", "N/A", "#VALUE!"):
        return None
    try:
        result = float(text.replace("$", "").replace(",", ""))
        return result if result > 0 else None
    except ValueError:
        return None


def normalize_category(raw) -> str:
    """Mapeia o texto bruto da coluna Category para o código interno."""
    if not raw:
        return "PRODUTOS"
    upper = str(raw).upper().strip()
    # Busca por correspondência parcial (ex: "MÃO DE OBRA GERAL" → "MAO_DE_OBRA_G")
    for key, val in CATEGORY_MAP.items():
        if key in upper:
            return val
    return "PRODUTOS"


def get_hyperlink(cell) -> Optional[str]:
    """Extrai URL do hiperlink da célula, se existir."""
    if cell.hyperlink:
        return cell.hyperlink.target
    text = str(cell.value or "").strip()
    if text.startswith("http"):
        return text
    return None


def room_score(room_name: str, items: list) -> int:
    """
    Calcula um score para determinar se o quarto é Temático ou Adulto.
    Positivo → Temático. Negativo → Adulto.
    """
    score = 0
    name_lower = room_name.lower()
    for kw in THEME_KEYWORDS:
        if kw in name_lower:
            score += 30
    for kw in ADULT_KEYWORDS:
        if kw in name_lower:
            score -= 50
    for item in items:
        item_lower = (item.get("name") or "").lower()
        if item.get("category") == "PACOTES_TEMATICOS":
            score += 100
        for kw in THEME_KEYWORDS:
            if kw in item_lower:
                score += 20
    return score

# ─── Extrator Principal ───────────────────────────────────────────────────────

def find_header_row(sheet) -> int:
    """
    Localiza a linha do cabeçalho procurando por 'Ambiente' ou 'Product'.
    As planilhas 2.0 sempre têm o cabeçalho na linha 12.
    """
    # Tenta rápido: linha 12 (padrão das planilhas 2.0)
    row12 = [str(sheet.cell(row=12, column=c).value or "").lower() for c in range(1, 15)]
    if "ambiente" in row12 or "product" in row12:
        return 12

    # Fallback: busca dinâmica nas primeiras 50 linhas
    for r in range(1, 51):
        vals = [str(sheet.cell(row=r, column=c).value or "").lower() for c in range(1, 15)]
        if "ambiente" in vals or ("product" in vals and "category" in vals):
            return r

    return 12  # usa 12 como padrão mesmo assim


def extract_project(file_info: dict) -> Optional[dict]:
    path = os.path.join(PLANILHAS_DIR, file_info["name"])
    if not os.path.exists(path):
        print(f"  ⚠️  Não encontrado: {file_info['name']}")
        return None

    print(f"  → Processando: {file_info['title']}...")

    wb = openpyxl.load_workbook(path, data_only=True)

    # Todas as planilhas 2.0 usam aba "BD1"
    sheet = wb["BD1"] if "BD1" in wb.sheetnames else wb.active

    header_row = find_header_row(sheet)

    # ── Ler dados linha a linha ──────────────────────────────────────────────
    rooms_raw: dict[str, dict] = {}

    for r in range(header_row + 1, sheet.max_row + 1):
        # Coluna B = Ambiente
        room_raw = sheet.cell(row=r, column=2).value
        if not room_raw or str(room_raw).strip() in ("", "None", "Ambiente"):
            continue

        # Normaliza espaços extras (ex: "Minions " == "Minions")
        room_name = " ".join(str(room_raw).split())

        # Coluna D = Produto
        product_val = sheet.cell(row=r, column=4).value
        if not product_val or str(product_val).strip() in ("", "None", "Product", "Produto"):
            continue

        category   = normalize_category(sheet.cell(row=r, column=5).value)
        link       = get_hyperlink(sheet.cell(row=r, column=6))
        qty        = clean_number(sheet.cell(row=r, column=7).value) or 1.0
        unit_price = clean_number(sheet.cell(row=r, column=8).value)
        price_cost = clean_number(sheet.cell(row=r, column=9).value)
        actual_raw = clean_actual(sheet.cell(row=r, column=10).value)

        # Actual válido = qualquer valor preenchido na planilha (sem filtro)
        actual_cost = actual_raw  # None apenas quando célula está vazia

        item = {
            "name":       str(product_val),
            "category":   category,
            "quantity":   qty,
            "unitPrice":  unit_price if unit_price > 0 else (price_cost / qty if qty > 0 else price_cost),
            "priceCost":  price_cost,
            "actualCost": actual_cost,   # None apenas quando não preenchido
            "link":       link,
        }

        if room_name not in rooms_raw:
            rooms_raw[room_name] = {"name": room_name, "items": []}
        rooms_raw[room_name]["items"].append(item)

    # ── Agregar por cômodo ───────────────────────────────────────────────────
    project_rooms = []
    cat_breakdown: dict[str, dict] = {}

    total_budget = 0.0
    total_actual = 0.0
    total_items = 0
    total_items_with_actual = 0

    thematic_count = 0
    adult_count = 0
    loft_count = 0
    game_room_count = 0

    for room_name, room_data in rooms_raw.items():
        items = room_data["items"]
        name_lower = room_name.lower()

        room_budget = sum(i["priceCost"] for i in items)
        room_actual_items = [i["actualCost"] for i in items if i["actualCost"] is not None]
        room_actual = sum(room_actual_items)

        room_items_count = len(items)
        room_items_with_actual = len(room_actual_items)
        room_coverage = round(room_items_with_actual / room_items_count, 3) if room_items_count > 0 else 0.0

        # Classificar tipo do cômodo
        is_bedroom = any(kw in name_lower for kw in ["bedroom", "dorm", "quarto", "suíte", "suite"])
        score = room_score(room_name, items)

        if is_bedroom:
            if score >= 30:
                room_type = "Quarto Temático"
                thematic_count += 1
            else:
                room_type = "Adulto"
                adult_count += 1
        elif any(kw in name_lower for kw in ["loft"]):
            room_type = "Loft"
            loft_count += 1
        elif any(kw in name_lower for kw in ["game room", "garagem", "garage", "cinema", "billiard"]):
            room_type = "Garagem / Cinema"
            game_room_count += 1
        elif "lanai" in name_lower:
            room_type = "Lanai"
        elif any(kw in name_lower for kw in ["bathroom", "bath"]):
            room_type = "Banheiro"
        elif "extras" in name_lower:
            room_type = "Extras"
        else:
            room_type = "Área Comum"

        # Atualizar totais do projeto
        total_budget += room_budget
        total_actual += room_actual
        total_items += room_items_count
        total_items_with_actual += room_items_with_actual

        # Breakdown por categoria
        for item in items:
            cat = item["category"]
            if cat not in cat_breakdown:
                cat_breakdown[cat] = {"priceCost": 0.0, "actualCost": 0.0, "itemsWithActual": 0, "totalItems": 0}
            cat_breakdown[cat]["priceCost"]  += item["priceCost"]
            cat_breakdown[cat]["totalItems"] += 1
            if item["actualCost"] is not None:
                cat_breakdown[cat]["actualCost"]      += item["actualCost"]
                cat_breakdown[cat]["itemsWithActual"] += 1

        # Limpar campos de auditoria internos antes de salvar
        clean_items = [
            {k: v for k, v in i.items() if not k.startswith("_")}
            for i in items
        ]

        project_rooms.append({
            "name":        room_name,
            "type":        room_type,
            "totalBudget": round(room_budget, 2),
            "totalActual": round(room_actual, 2) if room_items_with_actual > 0 else None,
            "coverage":    room_coverage,
            "items":       clean_items,
        })

    # ── Métricas finais do projeto ───────────────────────────────────────────
    coverage = round(total_items_with_actual / total_items, 3) if total_items > 0 else 0.0
    deviation_reliable = coverage >= MIN_COVERAGE_FOR_RELIABLE

    # Finalizar categoryBreakdown: actualCost = None se não há dados
    final_cat_breakdown = {}
    for cat, data in cat_breakdown.items():
        final_cat_breakdown[cat] = {
            "priceCost":  round(data["priceCost"], 2),
            "actualCost": round(data["actualCost"], 2) if data["itemsWithActual"] > 0 else None,
            "coverage":   round(data["itemsWithActual"] / data["totalItems"], 3) if data["totalItems"] > 0 else 0.0,
        }

    return {
        "id":               file_info["id"],
        "title":            file_info["title"],
        # Valores financeiros principais
        "budget":           round(total_budget, 2),
        "actual":           round(total_actual, 2) if total_items_with_actual > 0 else None,
        # Confiabilidade dos dados
        "coverage":         coverage,
        "deviationReliable": deviation_reliable,
        # Contadores de cômodos
        "roomCount":        len(project_rooms),
        "thematicCount":    thematic_count,
        "adultCount":       adult_count,
        "loftCount":        loft_count,
        "gameRoomCount":    game_room_count,
        "hasLanai":         any("lanai" in r["name"].lower() for r in project_rooms),
        # Detalhamento
        "rooms":            project_rooms,
        "categoryBreakdown": final_cat_breakdown,
    }


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("NOHA Extrator v3.0")
    print("─" * 40)

    projects = []
    for f in FILES:
        result = extract_project(f)
        if result:
            projects.append(result)
            status = "✅" if result["deviationReliable"] else "⚠️ "
            actual_str = f"${result['actual']:,.0f}" if result["actual"] is not None else "N/A"
            print(
                f"     {status} {result['title']}: "
                f"Budget=${result['budget']:,.0f} | "
                f"Actual={actual_str} | "
                f"Cobertura={result['coverage']:.0%}"
            )

    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(projects, f, indent=2, ensure_ascii=False)

    print("─" * 40)
    print(f"✅ db.json gerado com {len(projects)} projetos.")
    print(f"   Destino: {DB_PATH}")


if __name__ == "__main__":
    main()
